import sys
import os
import openpyxl
from openpyxl import load_workbook

file_path = "C:\\Users\\Jacob Graver\\Desktop\\Training Log V3.1 test copy.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Main program function
def main():
    assembly_update()


# The function to update the Asembly blocks
def assembly_update():
# getting input from the user and ensuring that it is a number and converting it into an int 
    while True:
        value = input("What is the number of Assembly trainings? ")
        if value.isdigit():
            pass
        else:
            print('Please enter a number.')

    sheet.cell(row=7, column=2).value = value
    sheet.cell(row=8, column=2).value = value

wb.save('Training Log V3.1 test copy1.xlsx')
sys.exit()
